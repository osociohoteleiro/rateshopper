from flask import Blueprint, request, jsonify
from werkzeug.utils import secure_filename
from src.models import db, Tarifa, ImportacaoLog, Hotel
from datetime import datetime, date
import os
import re
from openpyxl import load_workbook

tarifa_bp = Blueprint('tarifa', __name__)

UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def is_empty(value):
    """Verifica se um valor está vazio ou é None"""
    return value is None or value == '' or str(value).strip() == ''

def parse_brazilian_date(date_str):
    """Converte data no formato brasileiro DD/MM/AAAA para objeto date"""
    try:
        if is_empty(date_str):
            return None
        
        # Se já é um objeto datetime, extrair apenas a data
        if isinstance(date_str, datetime):
            return date_str.date()
        
        # Se é string, tentar converter
        if isinstance(date_str, str):
            # Remover espaços
            date_str = date_str.strip()
            
            # Tentar formato DD/MM/AAAA
            if '/' in date_str:
                parts = date_str.split('/')
                if len(parts) == 3:
                    day, month, year = parts
                    return date(int(year), int(month), int(day))
            
            # Tentar formato DD-MM-AAAA
            if '-' in date_str:
                parts = date_str.split('-')
                if len(parts) == 3:
                    day, month, year = parts
                    return date(int(year), int(month), int(day))
        
        return None
    except:
        return None

def parse_brazilian_price(price_str):
    """Converte preço no formato brasileiro para float"""
    try:
        if is_empty(price_str):
            return None
        
        # Se já é número
        if isinstance(price_str, (int, float)):
            return float(price_str)
        
        # Se é string, limpar e converter
        if isinstance(price_str, str):
            # Remover espaços e símbolos de moeda
            price_str = price_str.strip().replace('R$', '').replace(' ', '')
            
            # Substituir vírgula por ponto (formato brasileiro)
            price_str = price_str.replace(',', '.')
            
            return float(price_str)
        
        return None
    except:
        return None

@tarifa_bp.route('/upload', methods=['POST'])
def upload_tarifas():
    try:
        # Verificar se hotel foi selecionado
        hotel_id = request.form.get('hotel_id')
        if not hotel_id:
            return jsonify({'error': 'Hotel deve ser selecionado'}), 400
        
        # Verificar se hotel existe
        hotel = Hotel.query.get(hotel_id)
        if not hotel:
            return jsonify({'error': 'Hotel não encontrado'}), 404
        
        # Verificar se arquivo foi enviado
        if 'file' not in request.files:
            return jsonify({'error': 'Nenhum arquivo enviado'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'Nenhum arquivo selecionado'}), 400
        
        if not allowed_file(file.filename):
            return jsonify({'error': 'Tipo de arquivo não permitido. Use .xlsx ou .xls'}), 400
        
        # Criar diretório de upload se não existir
        os.makedirs(UPLOAD_FOLDER, exist_ok=True)
        
        # Salvar arquivo
        filename = secure_filename(file.filename)
        filepath = os.path.join(UPLOAD_FOLDER, filename)
        file.save(filepath)
        
        # Criar log de importação
        log = ImportacaoLog(
            hotel_id=hotel_id,
            arquivo_nome=filename,
            status='processando'
        )
        db.session.add(log)
        db.session.commit()
        
        # Processar planilha
        try:
            # Ler planilha usando openpyxl (formato: 3 colunas sem cabeçalho)
            workbook = load_workbook(filepath)
            sheet = workbook.active
            
            # Converter dados para lista de dicionários
            data = []
            for row in sheet.iter_rows(values_only=True):
                if row[0] is not None:  # Pular linhas vazias
                    data.append({
                        'data_checkin': row[0] if len(row) > 0 else None,
                        'data_checkout': row[1] if len(row) > 1 else None,
                        'preco': row[2] if len(row) > 2 else None
                    })
            
            # Verificar se tem dados
            if not data:
                raise ValueError("Planilha não contém dados válidos")
            
            total_registros = len(data)
            registros_validos = 0
            registros_erro = 0
            erros_detalhes = []
            
            for index, row in enumerate(data):
                try:
                    # Processar datas
                    data_checkin = parse_brazilian_date(row['data_checkin'])
                    data_checkout = parse_brazilian_date(row['data_checkout'])
                    preco = parse_brazilian_price(row['preco'])
                    
                    # Validar dados
                    if not data_checkin:
                        erros_detalhes.append(f"Linha {index + 1}: Data check-in inválida")
                        registros_erro += 1
                        continue
                    
                    if not data_checkout:
                        erros_detalhes.append(f"Linha {index + 1}: Data check-out inválida")
                        registros_erro += 1
                        continue
                    
                    if not preco or preco <= 0:
                        erros_detalhes.append(f"Linha {index + 1}: Preço inválido")
                        registros_erro += 1
                        continue
                    
                    if data_checkout <= data_checkin:
                        erros_detalhes.append(f"Linha {index + 1}: Data check-out deve ser posterior ao check-in")
                        registros_erro += 1
                        continue
                    
                    # Criar tarifa
                    tarifa = Tarifa(
                        hotel_id=hotel_id,
                        data_checkin=data_checkin,
                        data_checkout=data_checkout,
                        preco=preco,
                        importacao_id=log.id
                    )
                    
                    db.session.add(tarifa)
                    registros_validos += 1
                    
                except Exception as e:
                    erros_detalhes.append(f"Linha {index + 1}: {str(e)}")
                    registros_erro += 1
            
            # Atualizar log
            log.total_registros = total_registros
            log.registros_validos = registros_validos
            log.registros_erro = registros_erro
            log.status = 'concluido' if registros_erro == 0 else 'concluido_com_erros'
            log.detalhes_erro = '\n'.join(erros_detalhes) if erros_detalhes else None
            
            db.session.commit()
            
            # Remover arquivo temporário
            os.remove(filepath)
            
            return jsonify({
                'message': 'Upload processado com sucesso',
                'total_registros': total_registros,
                'registros_validos': registros_validos,
                'registros_erro': registros_erro,
                'erros': erros_detalhes[:10] if erros_detalhes else []  # Primeiros 10 erros
            })
            
        except Exception as e:
            log.status = 'erro'
            log.detalhes_erro = str(e)
            db.session.commit()
            
            # Remover arquivo temporário
            if os.path.exists(filepath):
                os.remove(filepath)
            
            return jsonify({'error': f'Erro ao processar planilha: {str(e)}'}), 500
        
    except Exception as e:
        return jsonify({'error': f'Erro interno: {str(e)}'}), 500

@tarifa_bp.route('/estatisticas', methods=['GET'])
def get_estatisticas():
    try:
        total_tarifas = Tarifa.query.count()
        total_hoteis = Hotel.query.count()
        
        # Calcular total de concorrentes (relacionamentos entre hotéis)
        total_concorrentes = 0
        try:
            for hotel in Hotel.query.all():
                total_concorrentes += len(hotel.concorrentes)
        except:
            total_concorrentes = 0
        
        # Última importação
        ultima_importacao = None
        try:
            ultima_importacao_obj = ImportacaoLog.query.order_by(ImportacaoLog.created_at.desc()).first()
            if ultima_importacao_obj:
                ultima_importacao = ultima_importacao_obj.to_dict()
        except:
            pass
        
        return jsonify({
            'total_tarifas': total_tarifas,
            'total_hoteis': total_hoteis,
            'total_concorrentes': total_concorrentes,
            'ultima_importacao': ultima_importacao
        })
    except Exception as e:
        # Retornar dados padrão em caso de erro
        return jsonify({
            'total_tarifas': 0,
            'total_hoteis': 0,
            'total_concorrentes': 0,
            'ultima_importacao': None,
            'error': str(e)
        }), 200  # Retornar 200 para evitar erro no frontend

@tarifa_bp.route('/tarifas', methods=['GET'])
def get_tarifas():
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 10, type=int)
        hotel_id = request.args.get('hotel_id', type=int)
        
        query = Tarifa.query
        
        if hotel_id:
            query = query.filter_by(hotel_id=hotel_id)
        
        tarifas = query.order_by(Tarifa.created_at.desc()).paginate(
            page=page, per_page=per_page, error_out=False
        )
        
        return jsonify({
            'tarifas': [t.to_dict() for t in tarifas.items],
            'total': tarifas.total,
            'pages': tarifas.pages,
            'current_page': page
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

